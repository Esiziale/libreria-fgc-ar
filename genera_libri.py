import re
import os
from openpyxl import load_workbook
from collections import Counter

EXCEL_PATH = "Lista_libri.xlsx"
OUTPUT_DIR = "content/libreria"

def iniziali(nome):
    if not nome:
        return "Anonimo"
    parti = str(nome).strip().split()
    return "".join(p[0].upper() + "." for p in parti if p)

def slug(testo):
    testo = str(testo).lower().strip()
    testo = re.sub(r"[àáâã]", "a", testo)
    testo = re.sub(r"[èéêë]", "e", testo)
    testo = re.sub(r"[ìíîï]", "i", testo)
    testo = re.sub(r"[òóôõ]", "o", testo)
    testo = re.sub(r"[ùúûü]", "u", testo)
    testo = re.sub(r"[^a-z0-9\s-]", "", testo)
    testo = re.sub(r"\s+", "-", testo)
    testo = re.sub(r"-+", "-", testo)
    return testo[:60].strip("-")

wb = load_workbook(EXCEL_PATH, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))

os.makedirs(OUTPUT_DIR, exist_ok=True)

donatori_counter = Counter()
generati = 0

for row in rows[2:]:
    if row[0] is None:
        continue

    numero, titolo, autore, categoria, donatore, disponibile = row

    titolo = str(titolo).strip() if titolo else "Senza titolo"
    autore = str(autore).strip() if autore else ""
    categoria = str(categoria).strip() if categoria else "Varia"
    disponibile_bool = str(disponibile).strip().lower() in ("si", "sì", "yes", "true")
    donatore_ini = iniziali(donatore)

    donatori_counter[donatore_ini] += 1

    filename = f"{numero:03d}-{slug(titolo)}.md"
    filepath = os.path.join(OUTPUT_DIR, filename)

    contenuto = f"""---
title: "{titolo.replace('"', "'")}"
autore: "{autore.replace('"', "'")}"
numero_catalogo: {numero}
categoria: "{categoria}"
donatore: "{donatore_ini}"
disponibile: {str(disponibile_bool).lower()}
draft: false
---
"""

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(contenuto)

    generati += 1

print(f"Generati {generati} file in '{OUTPUT_DIR}/'")
print(f"\nDonatori (con iniziali):")
for ini, count in donatori_counter.most_common():
    print(f"  {ini}: {count} libri")
